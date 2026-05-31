"""
Stage 3 — USDA food flavonoid.xlsx → tmp_* 테이블 적재.

설계 결정:
- openpyxl(read_only)로 시트 단위 스트리밍 → 메모리 안전
- 산출물 보존을 위해 CSV로도 저장 (`scripts/_run/xlsx_csv/`)
- DB 적재는 Python(executemany)으로 직접 (Stage 2 학습: LOAD DATA는 RFC4180 한계)
- datetime → date 변환, 빈 셀 → NULL
- 각 단계 시간/행수 출력 (보고서용)

사용:
    python 05_extract_and_load_xlsx.py
"""
from __future__ import annotations

import argparse
import csv
import time
from datetime import date, datetime
from pathlib import Path

import openpyxl
import mysql.connector

DEFAULT_XLSX = Path("data/raw/USDA food flavonoid.xlsx")
DEFAULT_OUT_DIR = Path("scripts/_run/xlsx_csv")


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    # 문자열인 경우 'YYYY-MM-DD' 시도
    try:
        return datetime.strptime(str(v).strip(), "%Y-%m-%d").date()
    except Exception:
        return None


def to_int(v):
    if v is None or v == "":
        return None
    return int(v)


def to_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def to_decimal(v):
    if v is None or v == "":
        return None
    return float(v)


SHEET_CONFIG = {
    "FLAVDESC": {
        "table": "tmp_flavdesc",
        "cols":  ["nutrient_code", "flavonoid_description", "flavonoid_class", "tagname", "unit", "decimals"],
        "conv":  [to_int, to_str, to_str, to_str, to_str, to_int],
    },
    "MAINFOODDESC": {
        "table": "tmp_mainfooddesc",
        "cols":  ["food_code", "start_date", "end_date", "main_food_description"],
        "conv":  [to_int, to_date, to_date, to_str],
    },
    "FLAVVAL": {
        "table": "tmp_flavval",
        "cols":  ["food_code", "nutrient_code", "start_date", "end_date", "nutrient_value"],
        "conv":  [to_int, to_int, to_date, to_date, to_decimal],
    },
}


def process_sheet(ws, conf, cur, cnx, out_dir: Path) -> tuple[int, float]:
    table = conf["table"]
    cols  = conf["cols"]
    conv  = conf["conv"]

    cur.execute(f"DELETE FROM `{table}`")
    cnx.commit()

    sql = (f"INSERT INTO `{table}` ({', '.join('`'+c+'`' for c in cols)}) "
           f"VALUES ({', '.join(['%s']*len(cols))})")

    csv_path = out_dir / f"{ws.title.lower()}.csv"

    t0 = time.perf_counter()
    rows_buf = []
    BATCH = 5000
    total = 0
    raw_rows_iter = ws.iter_rows(values_only=True)
    header = next(raw_rows_iter)  # 헤더 스킵

    with open(csv_path, "w", encoding="utf-8", newline="") as fcsv:
        w = csv.writer(fcsv, quoting=csv.QUOTE_ALL, lineterminator="\n")
        w.writerow(cols)
        for row in raw_rows_iter:
            if all(v is None or v == "" for v in row):
                continue  # 빈 행 skip
            converted = tuple(fn(v) for fn, v in zip(conv, row))
            rows_buf.append(converted)
            w.writerow(["" if v is None else (v.isoformat() if isinstance(v, date) else v) for v in converted])
            if len(rows_buf) >= BATCH:
                cur.executemany(sql, rows_buf)
                cnx.commit()
                total += len(rows_buf)
                rows_buf.clear()

    if rows_buf:
        cur.executemany(sql, rows_buf)
        cnx.commit()
        total += len(rows_buf)

    cur.execute(f"SELECT COUNT(*) FROM `{table}`")
    (cnt,) = cur.fetchone()
    elapsed = time.perf_counter() - t0
    print(f"[{table}] inserted={total:,}  table_count={cnt:,}  elapsed={elapsed:.2f}s  csv={csv_path}")
    return cnt, elapsed


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--host", default="localhost")
    p.add_argument("--port", type=int, default=3306)
    p.add_argument("--user", default="root")
    p.add_argument("--password", required=True)
    p.add_argument("--db", default="usda_fdc")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = p.parse_args()
    args.out_dir.mkdir(parents=True, exist_ok=True)

    cnx = mysql.connector.connect(
        host=args.host, port=args.port, user=args.user, password=args.password,
        database=args.db, charset="utf8mb4", autocommit=False, use_pure=True,
    )
    cur = cnx.cursor()
    cur.execute("SET SESSION sql_mode=''")
    cur.execute("SET SESSION foreign_key_checks=0")
    cur.execute("SET SESSION unique_checks=0")

    print(f"[INFO] opening {args.xlsx}")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    results = {}
    for sheet_name, conf in SHEET_CONFIG.items():
        ws = wb[sheet_name]
        print(f"[INFO] processing sheet '{sheet_name}'  max_row={ws.max_row}")
        cnt, sec = process_sheet(ws, conf, cur, cnx, args.out_dir)
        results[sheet_name] = (cnt, sec)

    wb.close()
    cur.close()
    cnx.close()

    print("\n=== Summary ===")
    for k, (c, s) in results.items():
        print(f"  {k:14s}  rows={c:>10,}  time={s:7.2f}s")


if __name__ == "__main__":
    main()
